from fastapi import FastAPI, APIRouter, HTTPException, status, Depends, Header, File, UploadFile, Request
from fastapi.responses import StreamingResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import asyncio
from pathlib import Path
from typing import List, Optional
from datetime import datetime, timezone
import uuid
from openpyxl import Workbook, load_workbook
from io import BytesIO
import re
import unicodedata
import shutil
from pydantic import BaseModel as PydanticBaseModel

from models import (
    UserCreate, User, UserLogin, UserResponse,
    BrandCreate, Brand,
    ProductCreate, Product, ProductSpecification,
    CategoryCreate, Category,
    AddToCartRequest, Cart, CartItem,
    Wishlist,
    ReviewCreate, Review,
    OrderCreate, Order,
    DashboardStats,
    MenuItem, HeroBanner, ServiceAlbum, FAQ, ContactInfo, SettingsCreate, Settings,
    PageCreate, Page,
    ContactRequestCreate, ContactRequest,
    NewsletterSubscriptionCreate, NewsletterSubscription,
    InstallmentRequestCreate, InstallmentRequest,
    GiftCreate, Gift, GiftConditionCreate, GiftCondition, GiftLeadCreate, GiftLead
)
from auth_utils import verify_password, get_password_hash, create_access_token
from dependencies import get_current_user, get_current_admin_user
from telegram_notifier import send_telegram_message, format_order_message

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

security = HTTPBearer()

# Helper function to generate slug from text
def generate_slug(text: str) -> str:
    """Generate URL-friendly slug from text"""
    # Normalize unicode characters
    text = unicodedata.normalize('NFKD', text)
    # Convert to lowercase and remove non-alphanumeric chars
    text = re.sub(r'[^\w\s-]', '', text).strip().lower()
    # Replace spaces with hyphens
    text = re.sub(r'[-\s]+', '-', text)
    return text

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ==================== AUTHENTICATION ENDPOINTS ====================

@api_router.post("/auth/register", response_model=UserResponse)
async def register(user_data: UserCreate):
    """Register a new user"""
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Email already registered"
        )
    
    # Create new user
    user_dict = user_data.dict()
    hashed_password = get_password_hash(user_dict.pop("password"))
    
    new_user = User(**user_dict)
    user_to_save = new_user.dict()
    user_to_save["password"] = hashed_password
    
    await db.users.insert_one(user_to_save)
    
    return UserResponse(**new_user.dict())


@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    """Login user and return JWT token"""
    # Find user
    user = await db.users.find_one({"email": credentials.email})
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password"
        )
    
    # Verify password
    if not verify_password(credentials.password, user["password"]):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password"
        )
    
    # Create access token
    access_token = create_access_token(data={"sub": user["id"]})
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": UserResponse(
            id=user["id"],
            email=user["email"],
            firstName=user["firstName"],
            lastName=user["lastName"],
            role=user["role"]
        )
    }


@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(authorization: Optional[str] = Header(None)):
    """Get current user"""
    current_user = await get_current_user(authorization)
    return UserResponse(**current_user)


@api_router.put("/auth/me", response_model=UserResponse)
async def update_me(
    user_data: dict,
    authorization: Optional[str] = Header(None)
):
    """Update current user profile"""
    current_user = await get_current_user(authorization)
    
    # Fields that can be updated
    allowed_fields = ['firstName', 'lastName', 'phone', 'address', 'city', 'postalCode']
    update_data = {k: v for k, v in user_data.items() if k in allowed_fields}
    
    if update_data:
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": update_data}
        )
    
    # Get updated user
    updated_user = await db.users.find_one({"id": current_user["id"]}, {"_id": 0, "password": 0})
    return UserResponse(**updated_user)


# ==================== PRODUCT ENDPOINTS ====================

def _build_search_query(search: str) -> dict:
    """Build a Mongo query that matches every word in `search` (case-insensitive,
    in any order) against either `name` or `nameRu`.

    Example: search='robinet baie' matches a product called 'Robinet pentru baie'
    because both 'robinet' and 'baie' appear in the name.
    """
    if not search:
        return {}
    words = [w for w in re.split(r"\s+", search.strip()) if w]
    if not words:
        return {}
    # Each word must appear somewhere in name OR nameRu
    and_clauses = []
    for w in words:
        escaped = re.escape(w)
        and_clauses.append({
            "$or": [
                {"name": {"$regex": escaped, "$options": "i"}},
                {"nameRu": {"$regex": escaped, "$options": "i"}},
            ]
        })
    return {"$and": and_clauses}


def _merge_query(base: dict, extra: dict) -> dict:
    """Merge a search sub-query into an existing Mongo query, preserving any
    existing `$and` / `$or` keys correctly."""
    if not extra:
        return base
    if "$and" in base:
        base["$and"].extend(extra.get("$and", [extra]))
    else:
        base.update(extra)
    return base


@api_router.get("/products", response_model=List[Product])
async def get_products(
    category: Optional[str] = None,
    search: Optional[str] = None,
    minPrice: Optional[float] = None,
    maxPrice: Optional[float] = None,
    brandId: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all products with optional filtering"""
    query = {}
    
    if category:
        # Match either the legacy single `category` field or the new `categories` array
        query["$or"] = [
            {"category": category},
            {"categories": category}
        ]
    
    if search:
        _merge_query(query, _build_search_query(search))
    
    if minPrice is not None or maxPrice is not None:
        query["price"] = {}
        if minPrice is not None:
            query["price"]["$gte"] = minPrice
        if maxPrice is not None:
            query["price"]["$lte"] = maxPrice
    
    if brandId:
        query["brandId"] = brandId
    
    # Cap limit defensively to avoid OOM in production
    limit = max(1, min(int(limit), 200))

    # Sort by createdAt descending (newest first)
    products = await db.products.find(query, {"_id": 0}).sort("createdAt", -1).skip(skip).limit(limit).to_list(limit)

    # Recompute the real review count and average rating for each product from the reviews collection
    product_ids = [p["id"] for p in products if p.get("id")]
    if product_ids:
        pipeline = [
            {"$match": {"productId": {"$in": product_ids}}},
            {"$group": {
                "_id": "$productId",
                "count": {"$sum": 1},
                "avgRating": {"$avg": "$rating"},
            }},
        ]
        stats = {
            s["_id"]: s async for s in db.reviews.aggregate(pipeline)
        }
        for p in products:
            s = stats.get(p.get("id"))
            if s:
                p["reviews"] = s["count"]
                p["rating"] = round(s["avgRating"], 1)
            else:
                p["reviews"] = 0
                p["rating"] = 0.0

    return [Product(**product) for product in products]


async def _distinct_brands_in_scope(scope_query: dict) -> list:
    """Return the list of distinct, non-empty brandId values for products
    matching the given scope query (typically a category filter)."""
    pipeline = [
        {"$match": scope_query} if scope_query else {"$match": {}},
        {"$group": {"_id": "$brandId"}},
    ]
    ids = []
    async for doc in db.products.aggregate(pipeline):
        bid = doc.get("_id")
        if bid:
            ids.append(bid)
    return ids


@api_router.get("/products/list/paginated")
async def get_products_paginated(
    category: Optional[str] = None,
    search: Optional[str] = None,
    minPrice: Optional[float] = None,
    maxPrice: Optional[float] = None,
    brandId: Optional[str] = None,
    brandIds: Optional[str] = None,  # comma-separated list of brand IDs
    page: int = 1,
    pageSize: int = 12,
):
    """Public paginated products endpoint: returns {items, total, page, pageSize, maxPrice}.

    Used by category/brand listing pages that need page-based navigation.
    """
    page = max(1, int(page))
    pageSize = max(1, min(int(pageSize), 100))

    query = {}
    if category:
        query["$or"] = [
            {"category": category},
            {"categories": category},
        ]
    if search:
        _merge_query(query, _build_search_query(search))
    if minPrice is not None or maxPrice is not None:
        price_q = {}
        if minPrice is not None:
            price_q["$gte"] = minPrice
        if maxPrice is not None:
            price_q["$lte"] = maxPrice
        query["price"] = price_q
    if brandIds:
        ids = [b.strip() for b in brandIds.split(",") if b.strip()]
        if ids:
            query["brandId"] = {"$in": ids}
    elif brandId:
        query["brandId"] = brandId

    # Compute the highest price within the unfiltered category scope (without
    # the price range filter) so the UI can show a usable range slider.
    scope_query = {}
    if category:
        scope_query["$or"] = [
            {"category": category},
            {"categories": category},
        ]
    max_doc = await db.products.find(scope_query, {"_id": 0, "price": 1}).sort("price", -1).limit(1).to_list(1)
    scope_max_price = float(max_doc[0]["price"]) if max_doc else 0

    total = await db.products.count_documents(query)
    skip = (page - 1) * pageSize
    cursor = (
        db.products.find(query, {"_id": 0})
        .sort("createdAt", -1)
        .skip(skip)
        .limit(pageSize)
    )
    items = await cursor.to_list(pageSize)

    # Recompute review counts (same as get_products)
    product_ids = [p.get("id") for p in items if p.get("id")]
    if product_ids:
        pipeline = [
            {"$match": {"productId": {"$in": product_ids}}},
            {"$group": {
                "_id": "$productId",
                "count": {"$sum": 1},
                "avgRating": {"$avg": "$rating"},
            }},
        ]
        stats = {s["_id"]: s async for s in db.reviews.aggregate(pipeline)}
        for p in items:
            s = stats.get(p.get("id"))
            p["reviews"] = s["count"] if s else 0
            p["rating"] = round(s["avgRating"], 1) if s else 0.0

    return {
        "items": [Product(**p).dict() for p in items],
        "total": total,
        "page": page,
        "pageSize": pageSize,
        "maxPrice": scope_max_price,
        "availableBrandIds": await _distinct_brands_in_scope(scope_query),
    }


@api_router.get("/admin/products")
async def admin_get_products(
    search: Optional[str] = None,
    page: int = 1,
    pageSize: int = 20,
    authorization: Optional[str] = Header(None)
):
    """Admin paginated products listing — returns {items, total, page, pageSize}."""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    page = max(1, int(page))
    pageSize = max(1, min(int(pageSize), 200))

    query = {}
    if search:
        _merge_query(query, _build_search_query(search))

    total = await db.products.count_documents(query)
    skip = (page - 1) * pageSize
    cursor = db.products.find(query, {"_id": 0}).sort("createdAt", -1).skip(skip).limit(pageSize)
    items = await cursor.to_list(pageSize)
    return {
        "items": [Product(**p).dict() for p in items],
        "total": total,
        "page": page,
        "pageSize": pageSize,
    }


@api_router.get("/products/{product_identifier}", response_model=Product)
async def get_product(product_identifier: str):
    """Get product by ID or slug"""
    # Try to find by slug first, then by ID
    product = await db.products.find_one({"slug": product_identifier}, {"_id": 0})
    
    if not product:
        # Fallback to ID for backward compatibility
        product = await db.products.find_one({"id": product_identifier}, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Recompute real review count and average rating from the reviews collection
    all_reviews = await db.reviews.find({"productId": product["id"]}, {"_id": 0}).to_list(1000)
    if all_reviews:
        product["reviews"] = len(all_reviews)
        product["rating"] = round(sum(r["rating"] for r in all_reviews) / len(all_reviews), 1)
    else:
        product["reviews"] = 0
        product["rating"] = 0.0

    return Product(**product)


@api_router.post("/products", response_model=Product)
async def create_product(
    product: ProductCreate,
    authorization: Optional[str] = Header(None)
):
    """Create a new product (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Generate slug if not provided
    product_dict = product.dict()
    if not product_dict.get("slug"):
        product_dict["slug"] = generate_slug(product.name)
    
    # Ensure slug is unique
    base_slug = product_dict["slug"]
    counter = 1
    while await db.products.find_one({"slug": product_dict["slug"]}):
        product_dict["slug"] = f"{base_slug}-{counter}"
        counter += 1
    
    new_product = Product(**product_dict)
    await db.products.insert_one(new_product.dict())
    
    # Update category item count
    await db.categories.update_one(
        {"name": product.category},
        {"$inc": {"itemCount": 1}}
    )
    
    return new_product


@api_router.put("/products/{product_id}", response_model=Product)
async def update_product(
    product_id: str,
    product: ProductCreate,
    authorization: Optional[str] = Header(None)
):
    """Update a product (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing_product = await db.products.find_one({"id": product_id})
    if not existing_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    updated_product = product.dict()
    
    # Generate slug if not provided or name changed
    if not updated_product.get("slug") or product.name != existing_product.get("name"):
        updated_product["slug"] = generate_slug(product.name)
        
        # Ensure slug is unique (exclude current product)
        base_slug = updated_product["slug"]
        counter = 1
        while True:
            existing_slug = await db.products.find_one({
                "slug": updated_product["slug"],
                "id": {"$ne": product_id}
            })
            if not existing_slug:
                break
            updated_product["slug"] = f"{base_slug}-{counter}"
            counter += 1
    
    updated_product["updatedAt"] = datetime.utcnow()
    
    await db.products.update_one(
        {"id": product_id},
        {"$set": updated_product}
    )
    
    updated_product["id"] = product_id
    updated_product["createdAt"] = existing_product["createdAt"]
    
    return Product(**updated_product)


@api_router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete a product (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if product exists
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Delete product
    await db.products.delete_one({"id": product_id})
    
    # Update category item count
    if product.get("category"):
        await db.categories.update_one(
            {"name": product["category"]},
            {"$inc": {"itemCount": -1}}
        )
    
    return {"message": "Product deleted successfully"}


# ==================== PRODUCTS IMPORT/EXPORT ENDPOINTS ====================

@api_router.get("/admin/products/export")
async def export_products(
    current_admin: dict = Depends(get_current_admin_user)
):
    """Export all products to Excel file (admin only).

    Exports every editable field from the Product model so the file can be
    re-imported without data loss. `categories` and `images` arrays are stored
    as comma-separated text; `specifications` as JSON.
    """
    import json as _json
    try:
        # Hard cap to prevent OOM in production. If catalog grows beyond this,
        # paginate the export instead.
        products = await db.products.find({}, {"_id": 0}).sort("createdAt", -1).limit(10000).to_list(10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Products"

        headers = [
            "ID",
            "Name RO", "Name RU",
            "Slug",
            "Description RO", "Description RU",
            "Price", "Original Price", "Discount",
            "Category (primary)", "Categories (comma-separated)",
            "Brand ID",
            "Store Name RO", "Store Name RU",
            "SKU",
            "Badge RO", "Badge RU",
            "Image (primary)",
            "Images (comma-separated)",
            "Colors (comma-separated)",
            "Sizes (comma-separated)",
            "Specifications (JSON)",
            "Available",
            "In Stock",
            "Is Active",
            "Rating", "Reviews", "Sold",
        ]
        ws.append(headers)

        def _csv(list_val):
            return ",".join([str(v) for v in (list_val or [])])

        for p in products:
            specs = p.get("specifications") or []
            # Serialize specs to JSON (safe, human-readable)
            specs_json = _json.dumps(specs, ensure_ascii=False) if specs else ""

            row = [
                p.get("id", ""),
                p.get("name", ""),
                p.get("nameRu", ""),
                p.get("slug", ""),
                p.get("description", ""),
                p.get("descriptionRu", ""),
                p.get("price", 0),
                p.get("originalPrice", ""),
                p.get("discount", 0),
                p.get("category", ""),
                _csv(p.get("categories")),
                p.get("brandId", "") or "",
                p.get("storeName", ""),
                p.get("storeNameRu", ""),
                p.get("sku", ""),
                p.get("badge", ""),
                p.get("badgeRu", ""),
                p.get("image", ""),
                _csv(p.get("images")),
                _csv(p.get("colors")),
                _csv(p.get("sizes")),
                specs_json,
                p.get("available", 0),
                bool(p.get("inStock", True)),
                bool(p.get("isActive", True)),
                p.get("rating", 0),
                p.get("reviews", 0),
                p.get("sold", 0),
            ]
            ws.append(row)

        # Widen columns a bit for readability
        for col_idx, _ in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=products.xlsx"}
        )

    except Exception as e:
        logger.error(f"Error exporting products: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Nu s-au putut exporta produsele: {str(e)}"
        )


@api_router.post("/admin/products/import")
async def import_products(
    file: UploadFile = File(...),
    current_admin: dict = Depends(get_current_admin_user)
):
    """Import products from Excel file (admin only).

    Column order must match the export format. Rows with an empty ID are
    treated as new products (UUID auto-generated); rows with an ID that
    already exists in the DB are updated in place.
    """
    import json as _json
    try:
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active

        rows = list(ws.iter_rows(min_row=2, values_only=True))

        def _csv_to_list(val):
            if val is None or val == "":
                return []
            return [v.strip() for v in str(val).split(",") if v.strip()]

        def _to_float(val, default=None):
            try:
                if val in (None, ""):
                    return default
                return float(val)
            except (TypeError, ValueError):
                return default

        def _to_int(val, default=0):
            try:
                if val in (None, ""):
                    return default
                return int(float(val))
            except (TypeError, ValueError):
                return default

        def _to_bool(val, default=True):
            if val is None or val == "":
                return default
            s = str(val).strip().lower()
            return s in ("true", "1", "yes", "y", "da", "t")

        def _col(row, idx, default=""):
            return row[idx] if len(row) > idx and row[idx] is not None else default

        imported_count = 0
        updated_count = 0
        errors = []

        for idx, row in enumerate(rows, start=2):
            try:
                # Skip fully-empty rows
                if all(c is None or c == "" for c in row):
                    continue
                # Require at minimum a Name RO
                name = str(_col(row, 1, "")).strip()
                if not name:
                    errors.append(f"Rând {idx}: lipsește Name RO")
                    continue

                raw_id = _col(row, 0, "")
                product_id = str(raw_id).strip() if raw_id else ""

                # Parse specifications (expect JSON; fall back to Python literal
                # for legacy exports)
                specs_raw = _col(row, 21, "")
                specifications = []
                if specs_raw:
                    try:
                        specifications = _json.loads(str(specs_raw))
                    except Exception:
                        try:
                            import ast
                            specifications = ast.literal_eval(str(specs_raw))
                        except Exception:
                            specifications = []
                if not isinstance(specifications, list):
                    specifications = []

                now = datetime.now(timezone.utc)
                product_data = {
                    "name": name,
                    "nameRu": str(_col(row, 2, "") or ""),
                    "slug": str(_col(row, 3, "") or ""),
                    "description": str(_col(row, 4, "") or ""),
                    "descriptionRu": str(_col(row, 5, "") or ""),
                    "price": _to_float(_col(row, 6), 0.0) or 0.0,
                    "originalPrice": _to_float(_col(row, 7), None),
                    "discount": _to_int(_col(row, 8), 0),
                    "category": str(_col(row, 9, "") or ""),
                    "categories": _csv_to_list(_col(row, 10)),
                    "brandId": (str(_col(row, 11, "")).strip() or None),
                    "storeName": str(_col(row, 12, "") or ""),
                    "storeNameRu": str(_col(row, 13, "") or ""),
                    "sku": str(_col(row, 14, "") or ""),
                    "badge": str(_col(row, 15, "") or ""),
                    "badgeRu": str(_col(row, 16, "") or ""),
                    "image": str(_col(row, 17, "") or ""),
                    "images": _csv_to_list(_col(row, 18)),
                    "colors": _csv_to_list(_col(row, 19)),
                    "sizes": _csv_to_list(_col(row, 20)),
                    "specifications": specifications,
                    "available": _to_int(_col(row, 22), 0),
                    "inStock": _to_bool(_col(row, 23), True),
                    "isActive": _to_bool(_col(row, 24), True),
                    "rating": _to_float(_col(row, 25), 0.0) or 0.0,
                    "reviews": _to_int(_col(row, 26), 0),
                    "sold": _to_int(_col(row, 27), 0),
                    "updatedAt": now,
                }

                if product_id:
                    existing = await db.products.find_one({"id": product_id})
                else:
                    existing = None

                if existing:
                    await db.products.update_one(
                        {"id": product_id},
                        {"$set": product_data}
                    )
                    updated_count += 1
                else:
                    # New product: generate UUID if missing and set createdAt
                    product_data["id"] = product_id or str(uuid.uuid4())
                    product_data["createdAt"] = now
                    await db.products.insert_one(product_data)
                    imported_count += 1

            except Exception as e:
                errors.append(f"Rând {idx}: {str(e)}")
                continue

        return {
            "message": (
                f"Import finalizat! {imported_count} produse noi, "
                f"{updated_count} actualizate"
                + (f", {len(errors)} erori" if errors else "")
            ),
            "imported": imported_count,
            "updated": updated_count,
            "errors": errors[:20],
        }

    except Exception as e:
        logger.error(f"Error importing products: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Nu s-au putut importa produsele: {str(e)}"
        )

async def delete_product(
    product_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete a product (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    await db.products.delete_one({"id": product_id})
    
    # Update category item count
    await db.categories.update_one(
        {"name": product["category"]},
        {"$inc": {"itemCount": -1}}
    )
    
    return {"message": "Product deleted successfully"}


class BulkDeleteRequest(PydanticBaseModel):
    ids: List[str]


@api_router.post("/admin/products/bulk-delete")
async def bulk_delete_products(
    payload: BulkDeleteRequest,
    current_admin: dict = Depends(get_current_admin_user)
):
    """Delete multiple products at once (admin only)."""
    ids = [i for i in (payload.ids or []) if i]
    if not ids:
        return {"deleted": 0, "message": "Nicio selecție"}

    result = await db.products.delete_many({"id": {"$in": ids}})
    return {
        "deleted": int(result.deleted_count),
        "message": f"{result.deleted_count} produse șterse"
    }


# ==================== REVIEWS ENDPOINTS ====================

@api_router.post("/products/{product_id}/reviews", response_model=Review)
async def create_review(product_id: str, review: ReviewCreate):
    """Create a review for a product"""
    # Check if product exists
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if user already reviewed this product (based on email)
    existing_review = await db.reviews.find_one({
        "productId": product_id,
        "userEmail": review.userEmail
    })
    if existing_review:
        raise HTTPException(status_code=400, detail="Ai lăsat deja o recenzie pentru acest produs")
    
    # Validate rating
    if review.rating < 1 or review.rating > 5:
        raise HTTPException(status_code=400, detail="Rating must be between 1 and 5")
    
    # Create review
    new_review = Review(**review.dict())
    await db.reviews.insert_one(new_review.dict())
    
    # Recalculate product rating
    all_reviews = await db.reviews.find({"productId": product_id}, {"_id": 0}).to_list(1000)
    avg_rating = sum(r["rating"] for r in all_reviews) / len(all_reviews)
    
    # Update product rating and review count
    await db.products.update_one(
        {"id": product_id},
        {"$set": {"rating": round(avg_rating, 1), "reviews": len(all_reviews)}}
    )
    
    return new_review


@api_router.get("/products/{product_id}/reviews", response_model=List[Review])
async def get_product_reviews(product_id: str):
    """Get all reviews for a product"""
    reviews = await db.reviews.find({"productId": product_id}, {"_id": 0}).sort("createdAt", -1).to_list(100)
    return [Review(**review) for review in reviews]


# ==================== CATEGORY ENDPOINTS ====================

@api_router.get("/categories", response_model=List[Category])
async def get_categories():
    """Get all categories with product count (single aggregation query)."""
    categories = await db.categories.find({}, {"_id": 0}).to_list(500)

    # Single aggregation: count products per category, taking into account
    # both the legacy `category` field and the new `categories` array.
    pipeline = [
        {
            "$project": {
                "names": {
                    "$concatArrays": [
                        [{"$ifNull": ["$category", ""]}],
                        {"$ifNull": ["$categories", []]},
                    ]
                }
            }
        },
        {"$unwind": "$names"},
        {"$match": {"names": {"$ne": ""}}},
        {"$group": {"_id": "$names", "count": {"$sum": 1}}},
    ]
    counts = {c["_id"]: c["count"] async for c in db.products.aggregate(pipeline)}

    for category in categories:
        category["itemCount"] = counts.get(category.get("name", ""), 0)

    return [Category(**cat) for cat in categories]


@api_router.get("/categories/{category_id}", response_model=Category)
async def get_category(category_id: str):
    """Get category by ID"""
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return Category(**category)


@api_router.post("/categories", response_model=Category)
async def create_category(
    category: CategoryCreate,
    authorization: Optional[str] = Header(None)
):
    """Create a new category (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if category already exists
    existing = await db.categories.find_one({"slug": category.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    
    new_category = Category(**category.dict())
    await db.categories.insert_one(new_category.dict())
    
    return new_category


@api_router.put("/categories/{category_id}", response_model=Category)
async def update_category(
    category_id: str,
    category: CategoryCreate,
    authorization: Optional[str] = Header(None)
):
    """Update a category (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing_category = await db.categories.find_one({"id": category_id})
    if not existing_category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    await db.categories.update_one(
        {"id": category_id},
        {"$set": category.dict()}
    )
    
    updated_category = category.dict()
    updated_category["id"] = category_id
    updated_category["createdAt"] = existing_category["createdAt"]
    
    return Category(**updated_category)


@api_router.delete("/categories/{category_id}")
async def delete_category(
    category_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete a category (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    category = await db.categories.find_one({"id": category_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    await db.categories.delete_one({"id": category_id})
    
    return {"message": "Category deleted successfully"}


# ==================== BRANDS ENDPOINTS ====================

@api_router.get("/brands", response_model=List[Brand])
async def get_brands():
    """Get all brands"""
    brands = await db.brands.find({}, {"_id": 0}).to_list(100)
    return [Brand(**brand) for brand in brands]


@api_router.get("/brands/{brand_id}", response_model=Brand)
async def get_brand(brand_id: str):
    """Get brand by ID"""
    brand = await db.brands.find_one({"id": brand_id}, {"_id": 0})
    if not brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    return Brand(**brand)


@api_router.post("/brands", response_model=Brand)
async def create_brand(
    brand: BrandCreate,
    authorization: Optional[str] = Header(None)
):
    """Create a new brand (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if brand already exists
    existing = await db.brands.find_one({"name": brand.name})
    if existing:
        raise HTTPException(status_code=400, detail="Brand already exists")
    
    new_brand = Brand(**brand.dict())
    await db.brands.insert_one(new_brand.dict())
    
    return new_brand


@api_router.put("/brands/{brand_id}", response_model=Brand)
async def update_brand(
    brand_id: str,
    brand: BrandCreate,
    authorization: Optional[str] = Header(None)
):
    """Update a brand (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing_brand = await db.brands.find_one({"id": brand_id})
    if not existing_brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    await db.brands.update_one(
        {"id": brand_id},
        {"$set": brand.dict()}
    )
    
    updated_brand = await db.brands.find_one({"id": brand_id}, {"_id": 0})
    return Brand(**updated_brand)


@api_router.delete("/brands/{brand_id}")
async def delete_brand(
    brand_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete a brand (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing_brand = await db.brands.find_one({"id": brand_id})
    if not existing_brand:
        raise HTTPException(status_code=404, detail="Brand not found")
    
    await db.brands.delete_one({"id": brand_id})
    
    return {"message": "Brand deleted successfully"}


# ==================== CART ENDPOINTS ====================

@api_router.get("/cart")
async def get_cart(authorization: Optional[str] = Header(None)):
    """Get user cart with populated product data"""
    current_user = await get_current_user(authorization)
    
    cart = await db.carts.find_one({"userId": current_user["id"]}, {"_id": 0})
    if not cart:
        # Create empty cart
        return {"userId": current_user["id"], "items": [], "total": 0.0}
    
    # Populate items with product data
    populated_items = []
    for item in cart.get("items", []):
        product = await db.products.find_one({"id": item["productId"]}, {"_id": 0})
        if product:
            populated_item = {
                **item,
                "id": product["id"],
                "name": product["name"],
                "image": product["image"],
                "images": product.get("images", []),
                "specifications": product.get("specifications", []),
                "category": product.get("category", ""),
                "description": product.get("description", "")
            }
            populated_items.append(populated_item)
    
    cart["items"] = populated_items
    return cart


@api_router.post("/cart/add")
async def add_to_cart(
    item: AddToCartRequest,
    authorization: Optional[str] = Header(None)
):
    """Add item to cart"""
    current_user = await get_current_user(authorization)
    
    # Get product
    product = await db.products.find_one({"id": item.productId})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Get or create cart
    cart = await db.carts.find_one({"userId": current_user["id"]})
    if not cart:
        cart = Cart(userId=current_user["id"]).dict()
        await db.carts.insert_one(cart)
    
    # Add item to cart
    cart_item = CartItem(
        productId=item.productId,
        quantity=item.quantity,
        selectedSize=item.selectedSize,
        selectedColor=item.selectedColor,
        price=product["price"]
    )
    
    # Check if item already exists
    items = cart.get("items", [])
    found = False
    for i, existing_item in enumerate(items):
        if (existing_item["productId"] == item.productId and
            existing_item.get("selectedSize") == item.selectedSize and
            existing_item.get("selectedColor") == item.selectedColor):
            items[i]["quantity"] += item.quantity
            found = True
            break
    
    if not found:
        items.append(cart_item.dict())
    
    # Calculate total
    total = sum(item["price"] * item["quantity"] for item in items)
    
    await db.carts.update_one(
        {"userId": current_user["id"]},
        {"$set": {"items": items, "total": total, "updatedAt": datetime.utcnow()}}
    )
    
    return {"message": "Item added to cart"}


@api_router.put("/cart/update")
async def update_cart_item(
    productId: str,
    quantity: int,
    selectedSize: Optional[str] = None,
    selectedColor: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """Update cart item quantity"""
    current_user = await get_current_user(authorization)
    
    cart = await db.carts.find_one({"userId": current_user["id"]})
    if not cart:
        raise HTTPException(status_code=404, detail="Cart not found")
    
    items = cart.get("items", [])
    for item in items:
        if (item["productId"] == productId and
            item.get("selectedSize") == selectedSize and
            item.get("selectedColor") == selectedColor):
            item["quantity"] = quantity
            break
    
    # Calculate total
    total = sum(item["price"] * item["quantity"] for item in items)
    
    await db.carts.update_one(
        {"userId": current_user["id"]},
        {"$set": {"items": items, "total": total, "updatedAt": datetime.utcnow()}}
    )
    
    return {"message": "Cart updated"}


@api_router.delete("/cart/remove/{product_id}")
async def remove_from_cart(
    product_id: str,
    selectedSize: Optional[str] = None,
    selectedColor: Optional[str] = None,
    authorization: Optional[str] = Header(None)
):
    """Remove item from cart"""
    current_user = await get_current_user(authorization)
    
    cart = await db.carts.find_one({"userId": current_user["id"]})
    if not cart:
        raise HTTPException(status_code=404, detail="Cart not found")
    
    items = cart.get("items", [])
    items = [
        item for item in items
        if not (item["productId"] == product_id and
                item.get("selectedSize") == selectedSize and
                item.get("selectedColor") == selectedColor)
    ]
    
    # Calculate total
    total = sum(item["price"] * item["quantity"] for item in items)
    
    await db.carts.update_one(
        {"userId": current_user["id"]},
        {"$set": {"items": items, "total": total, "updatedAt": datetime.utcnow()}}
    )
    
    return {"message": "Item removed from cart"}


@api_router.delete("/cart/clear")
async def clear_cart(authorization: Optional[str] = Header(None)):
    """Clear cart"""
    current_user = await get_current_user(authorization)
    
    await db.carts.update_one(
        {"userId": current_user["id"]},
        {"$set": {"items": [], "total": 0.0, "updatedAt": datetime.utcnow()}}
    )
    
    return {"message": "Cart cleared"}


# ==================== WISHLIST ENDPOINTS ====================

@api_router.get("/wishlist", response_model=Wishlist)
async def get_wishlist(authorization: Optional[str] = Header(None)):
    """Get user wishlist"""
    current_user = await get_current_user(authorization)
    
    wishlist = await db.wishlists.find_one({"userId": current_user["id"]})
    if not wishlist:
        new_wishlist = Wishlist(userId=current_user["id"])
        await db.wishlists.insert_one(new_wishlist.dict())
        return new_wishlist
    
    return Wishlist(**wishlist)


@api_router.post("/wishlist/add/{product_id}")
async def add_to_wishlist(
    product_id: str,
    authorization: Optional[str] = Header(None)
):
    """Add product to wishlist"""
    current_user = await get_current_user(authorization)
    
    # Check if product exists
    product = await db.products.find_one({"id": product_id})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Get or create wishlist
    wishlist = await db.wishlists.find_one({"userId": current_user["id"]})
    if not wishlist:
        wishlist = Wishlist(userId=current_user["id"]).dict()
        await db.wishlists.insert_one(wishlist)
    
    # Add product to wishlist
    products = wishlist.get("products", [])
    if product_id not in products:
        products.append(product_id)
    
    await db.wishlists.update_one(
        {"userId": current_user["id"]},
        {"$set": {"products": products}}
    )
    
    return {"message": "Product added to wishlist"}


@api_router.delete("/wishlist/remove/{product_id}")
async def remove_from_wishlist(
    product_id: str,
    authorization: Optional[str] = Header(None)
):
    """Remove product from wishlist"""
    current_user = await get_current_user(authorization)
    
    wishlist = await db.wishlists.find_one({"userId": current_user["id"]})
    if not wishlist:
        raise HTTPException(status_code=404, detail="Wishlist not found")
    
    products = wishlist.get("products", [])
    if product_id in products:
        products.remove(product_id)
    
    await db.wishlists.update_one(
        {"userId": current_user["id"]},
        {"$set": {"products": products}}
    )
    
    return {"message": "Product removed from wishlist"}


# ==================== ORDER ENDPOINTS ====================

@api_router.get("/orders", response_model=List[Order])
async def get_user_orders(authorization: Optional[str] = Header(None)):
    """Get user orders"""
    current_user = await get_current_user(authorization)
    
    orders = await db.orders.find({"userId": current_user["id"]}).to_list(100)
    return [Order(**order) for order in orders]


@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(
    order_id: str,
    authorization: Optional[str] = Header(None)
):
    """Get order by ID"""
    current_user = await get_current_user(authorization)
    
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Check if user owns the order or is admin
    if order["userId"] != current_user["id"] and current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    return Order(**order)


@api_router.post("/orders", response_model=Order)
async def create_order(order_data: OrderCreate):
    """Create a new order (guest or authenticated)"""
    # Create order
    new_order = Order(**order_data.dict())
    
    await db.orders.insert_one(new_order.dict(exclude={'_id'}))
    
    # Update product sold count and available stock
    for item in order_data.items:
        await db.products.update_one(
            {"id": item.productId},
            {
                "$inc": {
                    "sold": item.quantity,
                    "available": -item.quantity
                }
            }
        )
    
    # Fire-and-forget Telegram notification (never block or fail the order)
    try:
        asyncio.create_task(send_telegram_message(format_order_message(new_order)))
    except Exception as exc:
        logging.warning("Could not schedule Telegram notification: %s", exc)

    return new_order


# ==================== ADMIN ORDER ENDPOINTS ====================

@api_router.get("/admin/orders", response_model=List[Order])
async def get_all_orders(authorization: Optional[str] = Header(None)):
    """Get the most recent orders (admin only). Cap at 200 newest by default."""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    orders = await db.orders.find({}, {"_id": 0}).sort("createdAt", -1).limit(200).to_list(200)
    return [Order(**order) for order in orders]


@api_router.put("/admin/orders/{order_id}/status")
async def update_order_status(
    order_id: str,
    status: str,
    authorization: Optional[str] = Header(None)
):
    """Update order status (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"status": status, "updatedAt": datetime.utcnow()}}
    )
    
    return {"message": "Order status updated"}


# ==================== ADMIN USER ENDPOINTS ====================

@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_all_users(authorization: Optional[str] = Header(None)):
    """Get all users (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    users = await db.users.find().to_list(1000)
    return [UserResponse(**user) for user in users]


@api_router.get("/admin/users/{user_id}", response_model=UserResponse)
async def get_user(
    user_id: str,
    authorization: Optional[str] = Header(None)
):
    """Get user by ID (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return UserResponse(**user)


@api_router.delete("/admin/users/{user_id}")
async def delete_user(
    user_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete user (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.delete_one({"id": user_id})
    
    return {"message": "User deleted successfully"}


# ==================== ADMIN DASHBOARD ENDPOINTS ====================

@api_router.get("/admin/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats(authorization: Optional[str] = Header(None)):
    """Get dashboard statistics (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    total_users = await db.users.count_documents({})
    total_products = await db.products.count_documents({})
    total_orders = await db.orders.count_documents({})
    pending_orders = await db.orders.count_documents({"status": "pending"})
    low_stock_products = await db.products.count_documents({"available": {"$lt": 10}})
    
    # Calculate total revenue via MongoDB aggregation (no in-memory load)
    revenue_pipeline = [
        {"$group": {
            "_id": None,
            "total": {"$sum": {"$ifNull": ["$totalAmount", "$total"]}}
        }}
    ]
    revenue_result = await db.orders.aggregate(revenue_pipeline).to_list(1)
    total_revenue = float(revenue_result[0]["total"]) if revenue_result else 0.0

    return DashboardStats(
        totalUsers=total_users,
        totalProducts=total_products,
        totalOrders=total_orders,
        totalRevenue=total_revenue,
        pendingOrders=pending_orders,
        lowStockProducts=low_stock_products
    )


# ==================== SETTINGS ENDPOINTS ====================

@api_router.get("/settings")
async def get_settings():
    """Get site settings (menu configuration)"""
    settings = await db.settings.find_one({}, {"_id": 0})
    if not settings:
        # Return default empty settings
        default_settings = Settings(
            menuItems=[],
            categoryMenuItems=[]
        )
        return default_settings
    return Settings(**settings)


@api_router.post("/settings")
async def save_settings(
    settings_data: dict,  # Accept raw dict instead of Pydantic model
    authorization: Optional[str] = Header(None)
):
    """Save site settings (admin only) - Updates only provided fields"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if settings exist
    existing_settings = await db.settings.find_one({})
    
    # Prepare update with only provided fields
    update_fields = {**settings_data}  # Copy all provided fields
    update_fields["updatedAt"] = datetime.utcnow()
    
    if existing_settings:
        # Update existing - only set fields that were actually provided
        settings_id = existing_settings.get("id", str(uuid.uuid4()))
        update_fields["id"] = settings_id
        
        # Remove None values but keep empty lists/strings if explicitly provided
        clean_fields = {k: v for k, v in update_fields.items() if v is not None}
        
        await db.settings.update_one(
            {"id": settings_id},
            {"$set": clean_fields}
        )
    else:
        # Create new with provided fields + defaults
        update_fields["id"] = str(uuid.uuid4())
        
        # Set defaults for fields not provided
        defaults = {
            "menuItems": [],
            "categoryMenuItems": [],
            "featuredCategoryId": None,
            "heroBanners": [],
            "albums": [],
            "faqs": [],
            "contactInfo": {},
            "websiteName": "DOMIX",
            "favicon": ""
        }
        
        # Merge defaults with provided fields (provided fields take precedence)
        final_settings = {**defaults, **update_fields}
        
        await db.settings.insert_one(final_settings)
    
    return {"message": "Settings saved successfully"}


# ==================== PAGES ENDPOINTS ====================

@api_router.get("/pages", response_model=List[Page])
async def get_pages(published_only: bool = False):
    """Get all pages"""
    query = {}
    if published_only:
        query["isPublished"] = True
    
    pages = await db.pages.find(query, {"_id": 0}).to_list(100)
    return [Page(**page) for page in pages]


@api_router.get("/pages/{page_id}", response_model=Page)
async def get_page(page_id: str):
    """Get page by ID"""
    page = await db.pages.find_one({"id": page_id}, {"_id": 0})
    if not page:
        raise HTTPException(status_code=404, detail="Page not found")
    return Page(**page)


@api_router.get("/pages/slug/{slug}", response_model=Page)
async def get_page_by_slug(slug: str):
    """Get page by slug"""
    page = await db.pages.find_one({"slug": slug}, {"_id": 0})
    if not page:
        raise HTTPException(status_code=404, detail="Page not found")
    return Page(**page)


@api_router.post("/pages", response_model=Page)
async def create_page(
    page_data: PageCreate,
    authorization: Optional[str] = Header(None)
):
    """Create a new page (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if slug already exists
    existing = await db.pages.find_one({"slug": page_data.slug})
    if existing:
        raise HTTPException(status_code=400, detail="Page with this slug already exists")
    
    new_page = Page(**page_data.dict())
    await db.pages.insert_one(new_page.dict())
    
    return new_page


@api_router.put("/pages/{page_id}", response_model=Page)
async def update_page(
    page_id: str,
    page_data: PageCreate,
    authorization: Optional[str] = Header(None)
):
    """Update a page (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing_page = await db.pages.find_one({"id": page_id})
    if not existing_page:
        raise HTTPException(status_code=404, detail="Page not found")
    
    updated_page = page_data.dict()
    updated_page["updatedAt"] = datetime.utcnow()
    
    await db.pages.update_one(
        {"id": page_id},
        {"$set": updated_page}
    )
    
    updated_page["id"] = page_id
    updated_page["createdAt"] = existing_page["createdAt"]
    
    return Page(**updated_page)


@api_router.delete("/pages/{page_id}")
async def delete_page(
    page_id: str,
    authorization: Optional[str] = Header(None)
):
    """Delete a page (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    page = await db.pages.find_one({"id": page_id})
    if not page:
        raise HTTPException(status_code=404, detail="Page not found")
    
    await db.pages.delete_one({"id": page_id})
    
    return {"message": "Page deleted successfully"}


# ==================== CONTACT REQUESTS ENDPOINTS ====================

@api_router.post("/contact/submit")
async def submit_contact_request(request: ContactRequestCreate):
    """Submit a contact request"""
    try:
        contact_request = ContactRequest(**request.dict())
        await db.contact_requests.insert_one(contact_request.dict())
        return {"message": "Mesajul a fost trimis cu succes"}
    except Exception as e:
        logger.error(f"Error submitting contact request: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut trimite mesajul"
        )


@api_router.get("/admin/contact-requests")
async def get_contact_requests(
    current_admin: dict = Depends(get_current_admin_user)
):
    """Get all contact requests (admin only)"""
    try:
        requests = await db.contact_requests.find({}, {"_id": 0}).sort("createdAt", -1).to_list(1000)
        return requests
    except Exception as e:
        logger.error(f"Error fetching contact requests: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-au putut încărca solicitările"
        )


@api_router.put("/admin/contact-requests/{request_id}/status")
async def update_contact_request_status(
    request_id: str,
    status: str,
    current_admin: dict = Depends(get_current_admin_user)
):
    """Update contact request status (admin only)"""
    try:
        await db.contact_requests.update_one(
            {"id": request_id},
            {"$set": {"status": status}}
        )
        return {"message": "Status actualizat"}
    except Exception as e:
        logger.error(f"Error updating request status: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut actualiza statusul"
        )


@api_router.delete("/admin/contact-requests/{request_id}")
async def delete_contact_request(
    request_id: str,
    current_admin: dict = Depends(get_current_admin_user)
):
    """Delete contact request (admin only)"""
    try:
        await db.contact_requests.delete_one({"id": request_id})
        return {"message": "Solicitare ștearsă"}
    except Exception as e:
        logger.error(f"Error deleting contact request: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut șterge solicitarea"
        )


# ==================== NEWSLETTER ENDPOINTS ====================

@api_router.post("/newsletter/subscribe")
async def subscribe_to_newsletter(subscription: NewsletterSubscriptionCreate):
    """Subscribe to newsletter"""
    try:
        # Check if email already exists
        existing = await db.newsletter_subscriptions.find_one({"email": subscription.email})
        if existing:
            return {"message": "Adresa de email este deja abonată"}
        
        newsletter_sub = NewsletterSubscription(**subscription.dict())
        await db.newsletter_subscriptions.insert_one(newsletter_sub.dict())
        return {"message": "Abonare cu succes! Vei primi cele mai recente noutăți."}
    except Exception as e:
        logger.error(f"Error subscribing to newsletter: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut finaliza abonarea"
        )


@api_router.get("/admin/newsletter-subscriptions")
async def get_newsletter_subscriptions(
    current_admin: dict = Depends(get_current_admin_user)
):
    """Get all newsletter subscriptions (admin only)"""
    try:
        subscriptions = await db.newsletter_subscriptions.find({}, {"_id": 0}).sort("createdAt", -1).to_list(1000)
        return subscriptions
    except Exception as e:
        logger.error(f"Error fetching newsletter subscriptions: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-au putut încărca abonamentele"
        )


@api_router.delete("/admin/newsletter-subscriptions/{subscription_id}")
async def delete_newsletter_subscription(
    subscription_id: str,
    current_admin: dict = Depends(get_current_admin_user)
):
    """Delete newsletter subscription (admin only)"""
    try:
        await db.newsletter_subscriptions.delete_one({"id": subscription_id})
        return {"message": "Abonament șters"}
    except Exception as e:
        logger.error(f"Error deleting newsletter subscription: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut șterge abonamentul"
        )

    client.close()
@api_router.get("/")
async def root():
    return {"message": "Sellzy eCommerce API"}




# ==================== INSTALLMENT REQUESTS ENDPOINTS ====================

@api_router.post("/installment/request")
async def submit_installment_request(request: InstallmentRequestCreate):
    """Submit an installment payment request"""
    try:
        installment_request = InstallmentRequest(**request.dict())
        await db.installment_requests.insert_one(installment_request.dict())
        return {"message": "Cererea de plată în rate a fost trimisă cu succes"}
    except Exception as e:
        logger.error(f"Error submitting installment request: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut trimite cererea"
        )


@api_router.get("/admin/installment-requests")
async def get_installment_requests(
    current_admin: dict = Depends(get_current_admin_user)
):
    """Get all installment requests (admin only)"""
    try:
        requests = await db.installment_requests.find({}, {"_id": 0}).sort("createdAt", -1).to_list(1000)
        return requests
    except Exception as e:
        logger.error(f"Error fetching installment requests: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-au putut încărca cererile"
        )


@api_router.put("/admin/installment-requests/{request_id}/status")
async def update_installment_request_status(
    request_id: str,
    status: str,
    current_admin: dict = Depends(get_current_admin_user)
):
    """Update installment request status (admin only)"""
    try:
        await db.installment_requests.update_one(
            {"id": request_id},
            {"$set": {"status": status}}
        )
        return {"message": "Status actualizat"}
    except Exception as e:
        logger.error(f"Error updating installment request status: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut actualiza statusul"
        )


@api_router.delete("/admin/installment-requests/{request_id}")
async def delete_installment_request(
    request_id: str,
    current_admin: dict = Depends(get_current_admin_user)
):
    """Delete installment request (admin only)"""
    try:
        await db.installment_requests.delete_one({"id": request_id})
        return {"message": "Cerere ștearsă"}
    except Exception as e:
        logger.error(f"Error deleting installment request: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nu s-a putut șterge cererea"
        )

# ==================== FILE UPLOAD ENDPOINT ====================

@api_router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    authorization: Optional[str] = Header(None)
):
    """Upload image file (admin only)"""
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Validate file type
    allowed_types = ["image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"]
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400, 
            detail="Tipul fișierului nu este permis. Folosește: JPEG, PNG, WEBP, GIF"
        )
    
    # Validate file size (max 5MB)
    max_size = 5 * 1024 * 1024  # 5MB
    contents = await file.read()
    if len(contents) > max_size:
        raise HTTPException(
            status_code=400,
            detail="Fișierul este prea mare. Mărimea maximă: 5MB"
        )
    
    # Generate unique filename
    file_extension = file.filename.split('.')[-1]
    unique_filename = f"{uuid.uuid4()}.{file_extension}"
    
    # Save file to frontend/public/uploads
    upload_dir = Path("/app/frontend/public/uploads")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    file_path = upload_dir / unique_filename
    
    with open(file_path, "wb") as f:
        f.write(contents)
    
    # Return URL path (relative to public folder)
    file_url = f"/uploads/{unique_filename}"
    
    return {
        "url": file_url,
        "filename": unique_filename,
        "size": len(contents),
        "contentType": file.content_type
    }


# ==================== GIFT SYSTEM ENDPOINTS ====================

@api_router.get("/gifts", response_model=List[Gift])
async def get_gifts():
    """Get all gifts (public)"""
    gifts = await db.gifts.find({}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    return [Gift(**g) for g in gifts]


@api_router.post("/gifts", response_model=Gift)
async def create_gift(gift: GiftCreate, authorization: Optional[str] = Header(None)):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    new_gift = Gift(**gift.dict())
    await db.gifts.insert_one(new_gift.dict())
    return new_gift


@api_router.put("/gifts/{gift_id}", response_model=Gift)
async def update_gift(gift_id: str, gift: GiftCreate, authorization: Optional[str] = Header(None)):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    existing = await db.gifts.find_one({"id": gift_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Gift not found")
    await db.gifts.update_one({"id": gift_id}, {"$set": gift.dict()})
    updated = await db.gifts.find_one({"id": gift_id}, {"_id": 0})
    return Gift(**updated)


@api_router.delete("/gifts/{gift_id}")
async def delete_gift(gift_id: str, authorization: Optional[str] = Header(None)):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    await db.gifts.delete_one({"id": gift_id})
    return {"message": "Gift deleted"}


@api_router.get("/gift-conditions", response_model=List[GiftCondition])
async def get_gift_conditions():
    """Get all gift conditions (public – used by frontend popup matcher)"""
    conds = await db.gift_conditions.find({}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    return [GiftCondition(**c) for c in conds]


@api_router.post("/gift-conditions", response_model=GiftCondition)
async def create_gift_condition(
    cond: GiftConditionCreate,
    authorization: Optional[str] = Header(None)
):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    new_cond = GiftCondition(**cond.dict())
    await db.gift_conditions.insert_one(new_cond.dict())
    return new_cond


@api_router.put("/gift-conditions/{cond_id}", response_model=GiftCondition)
async def update_gift_condition(
    cond_id: str,
    cond: GiftConditionCreate,
    authorization: Optional[str] = Header(None)
):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    existing = await db.gift_conditions.find_one({"id": cond_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Condition not found")
    await db.gift_conditions.update_one({"id": cond_id}, {"$set": cond.dict()})
    updated = await db.gift_conditions.find_one({"id": cond_id}, {"_id": 0})
    return GiftCondition(**updated)


@api_router.delete("/gift-conditions/{cond_id}")
async def delete_gift_condition(cond_id: str, authorization: Optional[str] = Header(None)):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    await db.gift_conditions.delete_one({"id": cond_id})
    return {"message": "Condition deleted"}


@api_router.post("/gift-leads", response_model=GiftLead)
async def create_gift_lead(lead: GiftLeadCreate):
    """Public: capture lead from the gift popup (name + phone)."""
    new_lead = GiftLead(**lead.dict())
    await db.gift_leads.insert_one(new_lead.dict())

    # Notify via Telegram (non-blocking)
    try:
        gift_names = []
        if lead.giftIds:
            gifts = await db.gifts.find({"id": {"$in": lead.giftIds}}, {"_id": 0, "name": 1}).to_list(50)
            gift_names = [g.get("name", "") for g in gifts]
        msg = (
            "🎁 <b>Lead cadou nou</b>\n"
            "━━━━━━━━━━━━━━━━━\n"
            f"👤 <b>Nume:</b> {lead.customerName}\n"
            f"📞 <b>Telefon:</b> {lead.customerPhone}\n"
            f"🛍 <b>Produs:</b> {lead.productName or lead.productId}\n"
            f"🎁 <b>Cadouri:</b> {', '.join(gift_names) if gift_names else '—'}"
        )
        asyncio.create_task(send_telegram_message(msg))
    except Exception as exc:
        logging.warning("Could not send gift lead to Telegram: %s", exc)

    return new_lead


@api_router.get("/gift-leads", response_model=List[GiftLead])
async def get_gift_leads(authorization: Optional[str] = Header(None)):
    current_user = await get_current_user(authorization)
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    leads = await db.gift_leads.find({}, {"_id": 0}).sort("createdAt", -1).to_list(500)
    return [GiftLead(**l) for l in leads]


# ==================== SITEMAP ====================

def _sitemap_url_entry(loc: str, lastmod: Optional[str] = None,
                       changefreq: str = "weekly", priority: str = "0.7") -> str:
    from xml.sax.saxutils import escape
    parts = ["  <url>", f"    <loc>{escape(loc)}</loc>"]
    if lastmod:
        parts.append(f"    <lastmod>{escape(lastmod)}</lastmod>")
    parts.append(f"    <changefreq>{changefreq}</changefreq>")
    parts.append(f"    <priority>{priority}</priority>")
    parts.append("  </url>")
    return "\n".join(parts)


def _brand_slug(name: str) -> str:
    """Match the frontend BrandPage slug rule: lowercase + spaces→hyphens."""
    return re.sub(r"\s+", "-", (name or "").lower()).strip("-")


def _iso(dt) -> str:
    try:
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d")
        if isinstance(dt, str):
            return dt[:10]
    except Exception:
        pass
    return datetime.utcnow().strftime("%Y-%m-%d")


@api_router.get("/sitemap.xml")
async def sitemap(request: Request):
    """Dynamic XML sitemap for the store. Lists static pages, products,
    categories, brands and CMS pages."""
    # Prefer forwarded host/proto (set by the ingress / CDN) so URLs match the
    # public domain that search engines see, not the internal one.
    forwarded_host = request.headers.get("x-forwarded-host")
    forwarded_proto = request.headers.get("x-forwarded-proto")
    if forwarded_host:
        scheme = forwarded_proto or "https"
        base_url = f"{scheme}://{forwarded_host}"
    else:
        base_url = str(request.base_url).rstrip("/")
    # Allow explicit override via env var (useful in production)
    override = os.environ.get("PUBLIC_SITE_URL")
    if override:
        base_url = override.rstrip("/")
    today = datetime.utcnow().strftime("%Y-%m-%d")

    # 1) Static pages – keep in sync with App.js routes
    static_pages = [
        ("/", "daily", "1.0"),
        ("/catalog", "daily", "0.9"),
        ("/brands", "weekly", "0.7"),
        ("/servicii", "monthly", "0.6"),
        ("/despre-noi", "monthly", "0.5"),
        ("/intrebari-frecvente", "monthly", "0.5"),
        ("/contact", "monthly", "0.5"),
    ]

    entries: List[str] = []
    for path, cf, pr in static_pages:
        entries.append(_sitemap_url_entry(f"{base_url}{path}", today, cf, pr))

    # 2) Products
    products = await db.products.find(
        {}, {"_id": 0, "slug": 1, "id": 1, "updatedAt": 1, "createdAt": 1}
    ).to_list(5000)
    for p in products:
        ident = p.get("slug") or p.get("id")
        if not ident:
            continue
        lastmod = _iso(p.get("updatedAt") or p.get("createdAt"))
        entries.append(
            _sitemap_url_entry(f"{base_url}/product/{ident}", lastmod, "weekly", "0.8")
        )

    # 3) Categories
    categories = await db.categories.find({}, {"_id": 0, "slug": 1, "name": 1}).to_list(500)
    for c in categories:
        slug = c.get("slug") or c.get("name")
        if not slug:
            continue
        entries.append(
            _sitemap_url_entry(f"{base_url}/category/{slug}", today, "weekly", "0.7")
        )

    # 4) Brands (slug rule matches BrandPage.jsx)
    brands = await db.brands.find({}, {"_id": 0, "name": 1}).to_list(500)
    for b in brands:
        slug = _brand_slug(b.get("name", ""))
        if not slug:
            continue
        entries.append(
            _sitemap_url_entry(f"{base_url}/brand/{slug}", today, "weekly", "0.6")
        )

    # 5) CMS pages (published only)
    cms_pages = await db.pages.find(
        {"isPublished": True}, {"_id": 0, "slug": 1, "updatedAt": 1}
    ).to_list(500)
    for pg in cms_pages:
        slug = pg.get("slug")
        if not slug:
            continue
        lastmod = _iso(pg.get("updatedAt"))
        entries.append(
            _sitemap_url_entry(f"{base_url}/page/{slug}", lastmod, "monthly", "0.5")
        )

    xml_body = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        + "\n".join(entries)
        + "\n</urlset>\n"
    )
    return Response(content=xml_body, media_type="application/xml")


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_db_client():
    global db
    db = client[os.environ['DB_NAME']]
    logger.info("MongoDB connected")

@app.on_event("shutdown")
async def shutdown_db_client():
    logger.info("Shutting down")

